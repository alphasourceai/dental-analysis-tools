from __future__ import annotations

import csv
import json
import logging
import os
import re
import time
from datetime import datetime, timezone
from difflib import SequenceMatcher
from io import BytesIO, StringIO
from typing import Any, Callable, Optional

from supabase_utils import _get_supabase_admin_client

logger = logging.getLogger("uvicorn.error")

MAX_PROVIDER_RETRIES = 2
PROVIDER_RETRY_BACKOFF_SECONDS = 0.75
STORAGE_DOWNLOAD_MAX_RETRIES = 2
STORAGE_DOWNLOAD_RETRY_BACKOFF_SECONDS = 0.75
TRANSIENT_PROVIDER_STATUS_CODES = {408, 409, 425, 429, 500, 502, 503, 504, 529}
PROVIDER_UNAVAILABLE_MESSAGE = (
    "Analysis unavailable from this model due to temporary provider capacity. "
    "Other model results were processed."
)
PROVIDER_SKIPPED_MESSAGE = "Skipped for current admin xAI-only analysis mode."
MAX_XLSX_SHEETS = 10
MAX_XLSX_ROWS_PER_SHEET = 500
MAX_XLSX_COLUMNS_PER_ROW = 50
MAX_XLSX_CELL_CHARS = 500
MAX_PDF_PAGES = 30
MAX_PDF_CHARS = 60000
MAX_PDF_OCR_PAGES = 12
PDF_OCR_RENDER_SCALE = 2.0
STRUCTURED_ANALYSIS_SCHEMA_VERSION = "internal_analysis_v1"
STRUCTURED_ANALYSIS_START = "STRUCTURED_ANALYSIS_JSON_START"
STRUCTURED_ANALYSIS_END = "STRUCTURED_ANALYSIS_JSON_END"
MAX_STRUCTURED_FINDINGS = 12
MAX_STRUCTURED_LIST_ITEMS = 12
MAX_STRUCTURED_EVIDENCE_ITEMS = 5
MAX_STRUCTURED_TEXT_CHARS = 700
MAX_STRUCTURED_SHORT_TEXT_CHARS = 180
ADMIN_ANALYSIS_PROVIDER_MODE_ENV = "ADMIN_ANALYSIS_PROVIDER_MODE"
ADMIN_ANALYSIS_PROVIDER_MODE_XAI_ONLY = "xai_only"
ADMIN_ANALYSIS_PROVIDER_MODE_DEEP_REVIEW = "deep_review"
ADMIN_ANALYSIS_PROVIDER_MODE_DEFAULT = ADMIN_ANALYSIS_PROVIDER_MODE_XAI_ONLY
ADMIN_ANALYSIS_PROVIDER_ORDER = ("openai", "xai", "anthropic")

CancelChecker = Callable[[], bool]


class AdminFinancialProcessingError(Exception):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code
        self.message = message


class AdminFinancialProcessingCanceled(Exception):
    pass


def download_upload_file_bytes(bucket: str, object_path: str) -> bytes:
    if not bucket or not object_path:
        raise AdminFinancialProcessingError(
            "missing_storage_location",
            "Stored file location is missing.",
        )

    client = _get_supabase_admin_client()
    if not client:
        raise AdminFinancialProcessingError(
            "storage_not_configured",
            "File storage is not configured.",
        )

    max_attempts = STORAGE_DOWNLOAD_MAX_RETRIES + 1
    last_error: Optional[Exception] = None
    for attempt in range(1, max_attempts + 1):
        try:
            response = client.storage.from_(bucket).download(object_path)
            break
        except Exception as exc:
            last_error = exc
            logger.warning(
                "[admin_analysis] storage download failed bucket=%s object_path=%s attempt=%s error_type=%s",
                bucket,
                object_path,
                attempt,
                type(exc).__name__,
            )
            if attempt >= max_attempts or not _is_retryable_storage_download_error(exc):
                raise AdminFinancialProcessingError(
                    "storage_download_failed",
                    "Unable to download stored financial file.",
                ) from exc
            time.sleep(STORAGE_DOWNLOAD_RETRY_BACKOFF_SECONDS * attempt)
    else:
        logger.warning(
            "[admin_analysis] storage download failed bucket=%s object_path=%s attempt=%s error_type=%s",
            bucket,
            object_path,
            max_attempts,
            type(last_error).__name__ if last_error else "unknown",
        )
        raise AdminFinancialProcessingError(
            "storage_download_failed",
            "Unable to download stored financial file.",
        ) from last_error

    file_bytes = _response_to_bytes(response)
    if not file_bytes:
        raise AdminFinancialProcessingError(
            "storage_download_empty",
            "Stored financial file is empty.",
        )
    return file_bytes


def extract_csv_text(file_bytes: bytes) -> str:
    text = _decode_csv_bytes(file_bytes)
    rows = list(csv.reader(StringIO(text)))
    if not rows:
        raise AdminFinancialProcessingError(
            "empty_csv",
            "CSV file did not contain any rows.",
        )

    rendered_rows = []
    for row in rows:
        rendered_rows.append(" | ".join(str(value).strip() for value in row))
    data_input = "\n".join(rendered_rows).strip()
    if not data_input:
        raise AdminFinancialProcessingError(
            "empty_csv",
            "CSV file did not contain readable data.",
        )
    return data_input


def extract_xlsx_text(file_bytes: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise AdminFinancialProcessingError(
            "xlsx_dependency_missing",
            "XLSX processing is not configured.",
        ) from exc

    try:
        workbook = load_workbook(
            filename=BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise AdminFinancialProcessingError(
            "xlsx_read_failed",
            "XLSX file could not be read.",
        ) from exc

    rendered_sheets = []
    try:
        for sheet in workbook.worksheets[:MAX_XLSX_SHEETS]:
            rendered_rows = [f"Sheet: {_render_xlsx_cell(sheet.title)}"]
            for row_index, row in enumerate(
                sheet.iter_rows(
                    max_row=MAX_XLSX_ROWS_PER_SHEET,
                    max_col=MAX_XLSX_COLUMNS_PER_ROW,
                    values_only=True,
                ),
                start=1,
            ):
                rendered_cells = [_render_xlsx_cell(value) for value in row]
                while rendered_cells and not rendered_cells[-1]:
                    rendered_cells.pop()
                if not rendered_cells:
                    continue
                rendered_rows.append(" | ".join(rendered_cells))
                if row_index >= MAX_XLSX_ROWS_PER_SHEET:
                    break

            if len(rendered_rows) > 1:
                rendered_sheets.append("\n".join(rendered_rows))
    finally:
        workbook.close()

    data_input = "\n\n".join(rendered_sheets).strip()
    if not data_input:
        raise AdminFinancialProcessingError(
            "empty_xlsx",
            "XLSX file did not contain readable data.",
        )
    return data_input


def extract_pdf_text(file_bytes: bytes, *, enable_ocr: bool = False) -> str:
    try:
        import fitz
    except ImportError as exc:
        raise AdminFinancialProcessingError(
            "pdf_dependency_missing",
            "PDF processing is not configured.",
        ) from exc

    try:
        with fitz.open(stream=file_bytes, filetype="pdf") as document:
            if getattr(document, "needs_pass", False):
                raise AdminFinancialProcessingError(
                    "pdf_read_failed",
                    "PDF file could not be read.",
                )

            rendered_pages: list[str] = []
            extracted_chars = 0
            ocr_pages = 0
            ocr_dependencies: Optional[tuple[Any, Any]] = None
            for page_index, page in enumerate(document):
                if page_index >= MAX_PDF_PAGES or extracted_chars >= MAX_PDF_CHARS:
                    break

                page_text = _normalize_pdf_text(page.get_text("text") or "")
                if not page_text and enable_ocr and ocr_pages < MAX_PDF_OCR_PAGES:
                    if ocr_dependencies is None:
                        ocr_dependencies = _load_pdf_ocr_dependencies()
                    page_text = _extract_pdf_page_ocr_text(page, fitz, ocr_dependencies)
                    ocr_pages += 1

                if not page_text:
                    continue

                remaining_chars = MAX_PDF_CHARS - extracted_chars
                if len(page_text) > remaining_chars:
                    page_text = page_text[:remaining_chars].rstrip()

                rendered_pages.append(page_text)
                extracted_chars += len(page_text)
    except AdminFinancialProcessingError:
        raise
    except Exception as exc:
        raise AdminFinancialProcessingError(
            "pdf_extract_failed",
            "PDF text could not be extracted.",
        ) from exc

    data_input = "\n\n".join(rendered_pages).strip()
    if not data_input:
        if enable_ocr:
            raise AdminFinancialProcessingError(
                "empty_pdf_text",
                "PDF did not contain readable text after OCR. Upload a clearer PDF or a file with selectable text.",
            )
        raise AdminFinancialProcessingError(
            "empty_pdf_text",
            "PDF did not contain selectable text. Scanned or image-only PDFs are not supported yet because OCR is not enabled.",
        )
    return data_input


def run_financial_csv_analysis(
    data_input: str,
    *,
    cancel_checker: Optional[CancelChecker] = None,
    source_format: str = "csv",
    tool_type: str = "financial",
) -> dict[str, Any]:
    _raise_if_canceled(cancel_checker)
    normalized_tool_type = _normalize_tool_type(tool_type)
    model_labels = _get_model_labels()
    provider_specs = [
        ("openai", "OpenAI Analysis", "openai", _openai_analysis),
        ("xai", "xAI Analysis", "xai", _xai_analysis),
        ("anthropic", "AnthropicAI Analysis", "anthropic", _anthropic_analysis),
    ]
    enabled_providers = _get_admin_analysis_enabled_providers()

    raw_analyses: dict[str, str] = {}
    provider_statuses: dict[str, dict[str, Any]] = {}
    provider_structured_outputs: dict[str, Optional[dict[str, Any]]] = {}
    structured_provider_statuses: dict[str, dict[str, str]] = {}
    parsed_issues: dict[str, list[dict[str, Any]]] = {}
    parsed_trends: dict[str, list[dict[str, Any]]] = {}
    successful_provider_count = 0

    for provider_name, result_key, label_key, analysis_func in provider_specs:
        _raise_if_canceled(cancel_checker)
        if provider_name not in enabled_providers:
            raw_analyses[result_key] = PROVIDER_SKIPPED_MESSAGE
            provider_statuses[provider_name] = {
                "ok": False,
                "errorType": "skipped",
                "status": "skipped",
            }
            parsed_issues[label_key] = []
            parsed_trends[label_key] = []
            provider_structured_outputs[label_key] = None
            structured_provider_statuses[label_key] = {"status": "skipped"}
            continue

        analysis_text, succeeded, error_type = _run_provider_analysis_with_retry(
            provider_name=provider_name,
            analysis_func=lambda input_text, func=analysis_func: func(
                input_text,
                tool_type=normalized_tool_type,
            ),
            data_input=data_input,
        )
        _raise_if_canceled(cancel_checker)

        raw_analyses[result_key] = analysis_text
        provider_statuses[provider_name] = {
            "ok": succeeded,
            "errorType": error_type,
        }
        parsed_issues[label_key] = parse_issues_from_analysis(
            analysis_text,
            model_labels[label_key],
        )
        parsed_trends[label_key] = parse_trends_from_analysis(
            analysis_text,
            model_labels[label_key],
        )
        structured_output, structured_status = _parse_provider_structured_output(
            analysis_text,
            normalized_tool_type,
        )
        provider_structured_outputs[label_key] = structured_output
        structured_provider_statuses[label_key] = {"status": structured_status}
        if succeeded:
            successful_provider_count += 1

    if successful_provider_count == 0:
        raise AdminFinancialProcessingError(
            "provider_unavailable",
            "The analyzer providers are temporarily unavailable. Please try again later.",
        )

    all_issues = parsed_issues["openai"] + parsed_issues["xai"] + parsed_issues["anthropic"]
    all_trends = parsed_trends["openai"] + parsed_trends["xai"] + parsed_trends["anthropic"]
    deduplicated_issues = deduplicate_issues(all_issues)
    structured_analysis = _merge_structured_outputs(
        provider_structured_outputs,
        normalized_tool_type,
    )

    return {
        "sourceFormat": source_format,
        "toolType": normalized_tool_type,
        "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "raw_analyses": raw_analyses,
        "provider_statuses": provider_statuses,
        "provider_structured_outputs": provider_structured_outputs,
        "structured_provider_statuses": structured_provider_statuses,
        "structured_analysis": structured_analysis,
        "parsed_issues": parsed_issues,
        "parsed_trends": parsed_trends,
        "all_trends": all_trends,
        "deduplicated_issues": deduplicated_issues,
        "total_issue_count": len(deduplicated_issues),
    }


def parse_issues_from_analysis(analysis_text: str, source_model: str) -> list[dict[str, Any]]:
    issues = []

    if "---TRENDS---" in analysis_text:
        improvements_section = analysis_text.split("---TRENDS---")[0]
    else:
        improvements_section = analysis_text

    lines = improvements_section.strip().split("\n")
    current_issue: dict[str, Any] = {}

    for line in lines:
        line = line.strip()
        if not line:
            continue

        if line.upper().startswith("ISSUE:") or (len(line) > 0 and line[0].isdigit() and "." in line[:3]):
            if current_issue:
                issues.append(current_issue)

            issue_title = line.split(":", 1)[-1].strip() if ":" in line else line.split(".", 1)[-1].strip()
            current_issue = {
                "title": issue_title,
                "impact": "",
                "recommendation": "",
                "source": source_model,
                "full_text": line,
            }
        elif line.upper().startswith("IMPACT:"):
            if current_issue:
                current_issue["impact"] = line.split(":", 1)[-1].strip()
                current_issue["full_text"] += "\n" + line
        elif line.upper().startswith("RECOMMENDATION:"):
            if current_issue:
                current_issue["recommendation"] = line.split(":", 1)[-1].strip()
                current_issue["full_text"] += "\n" + line
        elif current_issue:
            current_issue["full_text"] += "\n" + line

    if current_issue:
        issues.append(current_issue)

    return issues


def parse_trends_from_analysis(analysis_text: str, source_model: str) -> list[dict[str, str]]:
    trends = []

    if "---TRENDS---" not in analysis_text:
        return trends

    trends_section = analysis_text.split("---TRENDS---")[1]
    lines = trends_section.strip().split("\n")

    for line in lines:
        line = line.strip()
        if not line:
            continue

        if "TREND:" in line.upper():
            trend_text = line.split(":", 1)[-1].strip() if ":" in line else line
            if len(trend_text) > 0 and trend_text[0].isdigit():
                trend_text = trend_text.split(".", 1)[-1].strip()

            if trend_text:
                trends.append(
                    {
                        "text": trend_text,
                        "source": source_model,
                    }
                )

    return trends


def deduplicate_issues(all_issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    def similar(a: str, b: str) -> bool:
        return SequenceMatcher(None, a.lower(), b.lower()).ratio() > 0.7

    deduplicated = []
    used_indices: set[int] = set()

    for index, issue in enumerate(all_issues):
        if index in used_indices:
            continue

        similar_issues = [issue]
        sources = [issue["source"]]

        for other_index, other_issue in enumerate(all_issues[index + 1 :], start=index + 1):
            if other_index in used_indices:
                continue

            if similar(issue["title"], other_issue["title"]):
                similar_issues.append(other_issue)
                sources.append(other_issue["source"])
                used_indices.add(other_index)

        deduplicated.append(
            {
                "title": issue["title"],
                "impact": issue["impact"],
                "recommendation": issue["recommendation"],
                "sources": sources,
                "count": len(sources),
                "all_versions": similar_issues,
            }
        )
        used_indices.add(index)

    return deduplicated


def _parse_provider_structured_output(
    analysis_text: str,
    tool_type: str,
) -> tuple[Optional[dict[str, Any]], str]:
    block = _extract_structured_json_block(analysis_text)
    if not block:
        return None, "missing"

    try:
        parsed = json.loads(block)
    except json.JSONDecodeError:
        return None, "invalid_json"

    if not isinstance(parsed, dict):
        return None, "validation_failed"

    normalized = _normalize_structured_analysis(parsed, tool_type)
    if not _structured_analysis_has_content(normalized):
        return None, "validation_failed"
    return normalized, "parsed"


def _extract_structured_json_block(analysis_text: str) -> str:
    if not isinstance(analysis_text, str) or not analysis_text:
        return ""

    start_index = analysis_text.find(STRUCTURED_ANALYSIS_START)
    end_index = analysis_text.find(STRUCTURED_ANALYSIS_END)
    if start_index < 0 or end_index < 0 or end_index <= start_index:
        return ""

    block = analysis_text[start_index + len(STRUCTURED_ANALYSIS_START) : end_index].strip()
    return _strip_json_fence(block)


def _strip_json_fence(value: str) -> str:
    text = (value or "").strip()
    if text.startswith("```"):
        lines = text.splitlines()
        if lines:
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        text = "\n".join(lines).strip()
    return text


def _normalize_structured_analysis(value: dict[str, Any], tool_type: str) -> dict[str, Any]:
    normalized_tool_type = _normalize_tool_type(value.get("toolType") or tool_type)
    return {
        "schemaVersion": STRUCTURED_ANALYSIS_SCHEMA_VERSION,
        "toolType": normalized_tool_type,
        "executiveSummary": _normalize_executive_summary(value.get("executiveSummary")),
        "rankedFindings": _normalize_ranked_findings(value.get("rankedFindings")),
        "dataQualityNotes": _normalize_structured_text_list(value.get("dataQualityNotes")),
        "implementationPriorities": _normalize_structured_text_list(value.get("implementationPriorities")),
        "consultantChecklist": _normalize_structured_text_list(value.get("consultantChecklist")),
        "suggestedReportSections": _normalize_structured_text_list(value.get("suggestedReportSections")),
    }


def _normalize_executive_summary(value: Any) -> dict[str, str]:
    source = value if isinstance(value, dict) else {}
    return {
        "summary": _structured_text(source.get("summary")),
        "primaryConcern": _structured_text(source.get("primaryConcern")),
        "recommendedFocus": _structured_text(source.get("recommendedFocus")),
    }


def _normalize_ranked_findings(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []

    findings: list[dict[str, Any]] = []
    for index, item in enumerate(value[:MAX_STRUCTURED_FINDINGS], start=1):
        if not isinstance(item, dict):
            continue
        title = _structured_text(item.get("title"), max_chars=MAX_STRUCTURED_SHORT_TEXT_CHARS)
        operational_implication = _structured_text(item.get("operationalImplication"))
        recommended_action = _structured_text(item.get("recommendedAction"))
        client_summary = _structured_text(item.get("clientFacingSummary"))
        if not any((title, operational_implication, recommended_action, client_summary)):
            continue

        findings.append(
            {
                "rank": _structured_rank(item.get("rank"), index),
                "title": title,
                "category": _structured_text(item.get("category"), max_chars=MAX_STRUCTURED_SHORT_TEXT_CHARS),
                "severity": _structured_choice(
                    item.get("severity"),
                    {"low", "medium", "high", "critical"},
                    "medium",
                ),
                "confidence": _structured_choice(
                    item.get("confidence"),
                    {"low", "medium", "high"},
                    "medium",
                ),
                "evidence": _normalize_structured_evidence(item.get("evidence")),
                "financialValue": _structured_text(item.get("financialValue"), max_chars=MAX_STRUCTURED_SHORT_TEXT_CHARS),
                "operationalImplication": operational_implication,
                "recommendedAction": recommended_action,
                "followUpQuestion": _structured_text(item.get("followUpQuestion")),
                "implementationDifficulty": _structured_choice(
                    item.get("implementationDifficulty"),
                    {"low", "medium", "high"},
                    "medium",
                ),
                "estimatedImpactCategory": _structured_choice(
                    item.get("estimatedImpactCategory"),
                    {
                        "cash_flow",
                        "revenue_leakage",
                        "workflow_efficiency",
                        "growth",
                        "compliance",
                        "data_quality",
                    },
                    "workflow_efficiency",
                ),
                "clientFacingSummary": client_summary,
                "internalReviewerNotes": _structured_text(item.get("internalReviewerNotes")),
            }
        )
    return findings


def _normalize_structured_evidence(value: Any) -> list[dict[str, str]]:
    if not isinstance(value, list):
        return []

    evidence_items: list[dict[str, str]] = []
    for item in value[:MAX_STRUCTURED_EVIDENCE_ITEMS]:
        if not isinstance(item, dict):
            continue
        evidence = {
            "label": _structured_text(item.get("label"), max_chars=MAX_STRUCTURED_SHORT_TEXT_CHARS),
            "value": _structured_text(item.get("value"), max_chars=MAX_STRUCTURED_SHORT_TEXT_CHARS),
            "sourceHint": _structured_text(item.get("sourceHint"), max_chars=MAX_STRUCTURED_SHORT_TEXT_CHARS),
        }
        if any(evidence.values()):
            evidence_items.append(evidence)
    return evidence_items


def _normalize_structured_text_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []

    items: list[str] = []
    seen: set[str] = set()
    for item in value:
        text = _structured_text(item)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        items.append(text)
        if len(items) >= MAX_STRUCTURED_LIST_ITEMS:
            break
    return items


def _structured_text(value: Any, *, max_chars: int = MAX_STRUCTURED_TEXT_CHARS) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return ""
    text = " ".join(str(value).strip().split())
    if not text:
        return ""
    if _structured_text_looks_sensitive(text):
        return ""
    if len(text) > max_chars:
        return f"{text[:max_chars].rstrip()}..."
    return text


def _structured_text_looks_sensitive(text: str) -> bool:
    lowered = text.lower()
    sensitive_markers = (
        "signed_url",
        "signed url",
        "token",
        "secret",
        "api key",
        "apikey",
        "password",
        "gcs path",
        "gs://",
        "storage/v1/object",
    )
    if any(marker in lowered for marker in sensitive_markers):
        return True
    if re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", text, flags=re.IGNORECASE):
        return True
    if re.search(r"\b\d{3}[-.\s]\d{2}[-.\s]\d{4}\b", text):
        return True
    if re.search(r"\b(?:\+?1[-.\s]?)?(?:\(?\d{3}\)?[-.\s]?)\d{3}[-.\s]?\d{4}\b", text):
        return True
    return False


def _structured_rank(value: Any, fallback: int) -> int:
    try:
        rank = int(value)
    except (TypeError, ValueError):
        return fallback
    return rank if rank > 0 else fallback


def _structured_choice(value: Any, allowed: set[str], default: str) -> str:
    normalized = str(value or "").strip().lower()
    return normalized if normalized in allowed else default


def _structured_analysis_has_content(value: dict[str, Any]) -> bool:
    summary = value.get("executiveSummary") or {}
    return bool(
        any(summary.get(key) for key in ("summary", "primaryConcern", "recommendedFocus"))
        or value.get("rankedFindings")
        or value.get("dataQualityNotes")
        or value.get("implementationPriorities")
        or value.get("consultantChecklist")
        or value.get("suggestedReportSections")
    )


def _merge_structured_outputs(
    provider_outputs: dict[str, Optional[dict[str, Any]]],
    tool_type: str,
) -> Optional[dict[str, Any]]:
    valid_outputs = [output for output in provider_outputs.values() if isinstance(output, dict)]
    if not valid_outputs:
        return None

    first_output = valid_outputs[0]
    merged = {
        "schemaVersion": STRUCTURED_ANALYSIS_SCHEMA_VERSION,
        "toolType": _normalize_tool_type(tool_type),
        "executiveSummary": first_output.get("executiveSummary")
        or _normalize_executive_summary({}),
        "rankedFindings": _merge_ranked_findings(valid_outputs),
        "dataQualityNotes": _merge_structured_text_lists(valid_outputs, "dataQualityNotes"),
        "implementationPriorities": _merge_structured_text_lists(valid_outputs, "implementationPriorities"),
        "consultantChecklist": _merge_structured_text_lists(valid_outputs, "consultantChecklist"),
        "suggestedReportSections": _merge_structured_text_lists(valid_outputs, "suggestedReportSections"),
    }
    return merged if _structured_analysis_has_content(merged) else None


def _merge_ranked_findings(outputs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    seen: set[str] = set()
    for output in outputs:
        findings = output.get("rankedFindings")
        if not isinstance(findings, list):
            continue
        for finding in findings:
            if not isinstance(finding, dict):
                continue
            title = (finding.get("title") or "").strip()
            key = title.lower() or json.dumps(finding, sort_keys=True, default=str)[:120]
            if key in seen:
                continue
            seen.add(key)
            copied = dict(finding)
            copied["rank"] = len(merged) + 1
            merged.append(copied)
            if len(merged) >= MAX_STRUCTURED_FINDINGS:
                return merged
    return merged


def _merge_structured_text_lists(outputs: list[dict[str, Any]], key: str) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for output in outputs:
        items = output.get(key)
        if not isinstance(items, list):
            continue
        for item in items:
            text = _structured_text(item)
            if not text:
                continue
            normalized = text.lower()
            if normalized in seen:
                continue
            seen.add(normalized)
            merged.append(text)
            if len(merged) >= MAX_STRUCTURED_LIST_ITEMS:
                return merged
    return merged


def _response_to_bytes(response: Any) -> bytes:
    if isinstance(response, bytes):
        return response
    if isinstance(response, bytearray):
        return bytes(response)
    if isinstance(response, str):
        return response.encode("utf-8")

    content = getattr(response, "content", None)
    if isinstance(content, bytes):
        return content
    data = getattr(response, "data", None)
    if isinstance(data, bytes):
        return data
    if isinstance(data, bytearray):
        return bytes(data)
    if isinstance(data, str):
        return data.encode("utf-8")

    if isinstance(response, dict):
        dict_data = response.get("data") or response.get("content")
        if isinstance(dict_data, bytes):
            return dict_data
        if isinstance(dict_data, bytearray):
            return bytes(dict_data)
        if isinstance(dict_data, str):
            return dict_data.encode("utf-8")

    return b""


def _decode_csv_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise AdminFinancialProcessingError(
        "csv_decode_failed",
        "CSV file could not be decoded.",
    )


def _is_retryable_storage_download_error(exc: Exception) -> bool:
    status_code = _provider_status_code(exc)
    if status_code is not None:
        return status_code in TRANSIENT_PROVIDER_STATUS_CODES

    safe_text = f"{type(exc).__name__} {str(exc)}".lower()
    non_retryable_markers = (
        "not found",
        "404",
        "unauthorized",
        "401",
        "forbidden",
        "403",
        "permission denied",
    )
    if any(marker in safe_text for marker in non_retryable_markers):
        return False

    retryable_markers = (
        "timeout",
        "timed out",
        "readtimeout",
        "connecttimeout",
        "connection",
        "network",
        "temporarily unavailable",
        "service unavailable",
        "bad gateway",
        "gateway timeout",
        "server error",
        "internal server",
        "rate limit",
        "408",
        "429",
        "500",
        "502",
        "503",
        "504",
        "529",
        "storage",
        "supabase",
    )
    return any(marker in safe_text for marker in retryable_markers)


def _render_xlsx_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        text = value.isoformat()
    else:
        text = str(value)
    text = " ".join(text.strip().split())
    if len(text) > MAX_XLSX_CELL_CHARS:
        return f"{text[:MAX_XLSX_CELL_CHARS]}..."
    return text


def _normalize_pdf_text(value: str) -> str:
    lines = [" ".join(line.strip().split()) for line in value.splitlines()]
    return "\n".join(line for line in lines if line)


def _load_pdf_ocr_dependencies() -> tuple[Any, Any]:
    try:
        from PIL import Image
        import pytesseract
    except ImportError as exc:
        raise AdminFinancialProcessingError(
            "pdf_ocr_dependency_missing",
            "PDF OCR processing is not configured.",
        ) from exc
    return Image, pytesseract


def _extract_pdf_page_ocr_text(
    page: Any,
    fitz_module: Any,
    ocr_dependencies: tuple[Any, Any],
) -> str:
    Image, pytesseract = ocr_dependencies
    try:
        matrix = fitz_module.Matrix(PDF_OCR_RENDER_SCALE, PDF_OCR_RENDER_SCALE)
        pixmap = page.get_pixmap(matrix=matrix, alpha=False)
        with Image.open(BytesIO(pixmap.tobytes("png"))) as image:
            return _normalize_pdf_text(pytesseract.image_to_string(image) or "")
    except Exception as exc:
        if type(exc).__name__ == "TesseractNotFoundError":
            raise AdminFinancialProcessingError(
                "pdf_ocr_dependency_missing",
                "PDF OCR processing is not configured.",
            ) from exc
        raise AdminFinancialProcessingError(
            "pdf_ocr_failed",
            "PDF OCR text could not be extracted.",
        ) from exc


def _get_model_config() -> dict[str, str]:
    return {
        "openai": os.getenv("OPENAI_MODEL", "gpt-5-chat-latest"),
        "xai": os.getenv("XAI_MODEL", "grok-4.3"),
        "anthropic": os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-5"),
    }


def _get_admin_analysis_enabled_providers() -> set[str]:
    mode = os.getenv(
        ADMIN_ANALYSIS_PROVIDER_MODE_ENV,
        ADMIN_ANALYSIS_PROVIDER_MODE_DEFAULT,
    ).strip().lower()
    if mode in {
        "multi_provider",
        "three_provider",
        "all",
        ADMIN_ANALYSIS_PROVIDER_MODE_DEEP_REVIEW,
    }:
        return set(ADMIN_ANALYSIS_PROVIDER_ORDER)
    return {"xai"}


def _get_model_labels() -> dict[str, str]:
    models = _get_model_config()
    return {
        "openai": f"OpenAI ({models.get('openai')})" if models.get("openai") else "OpenAI",
        "xai": f"xAI ({models.get('xai')})" if models.get("xai") else "xAI",
        "anthropic": (
            f"Anthropic ({models.get('anthropic')})"
            if models.get("anthropic")
            else "Anthropic"
        ),
    }


def _normalize_tool_type(tool_type: object) -> str:
    normalized = str(tool_type or "").strip().lower()
    if normalized in {"ar", "claims"}:
        return normalized
    return "financial"


def _get_analysis_prompt(tool_type: str = "financial") -> str:
    normalized_tool_type = _normalize_tool_type(tool_type)
    return f"""You are an expert dental operations consultant with deep knowledge of practice management, revenue cycle, and operational efficiency.

IMPORTANT FORMATTING RULES:
- Use PLAIN TEXT only - no LaTeX, no math formatting, no special markup
- Write dollar amounts as plain text: $10,000 not $10,000$ or escaped dollar amounts
- Do not use asterisks for emphasis or formatting
- Keep all text on single lines without special characters

Analyze the provided {normalized_tool_type.upper()} data and identify improvement opportunities AND key trends.

TOOL-SPECIFIC REVIEW FOCUS:
{_tool_specific_prompt_focus(normalized_tool_type)}

SECTION 1 - IMPROVEMENT OPPORTUNITIES:
Identify AT LEAST 3-5 high-level strategic areas for improvement. Format each as:
ISSUE: [Brief, strategic title - keep it high-level, not overly specific]
IMPACT: [Why this matters - financial, operational, or compliance impact]
RECOMMENDATION: [General recommended action]

Focus on strategic opportunities in:
- Revenue cycle optimization
- Operational efficiency
- Cost management
- Patient experience
- Compliance and risk management
- Technology utilization
- Staff productivity

SECTION 2 - KEY TRENDS:
After the improvement opportunities, add a separator line "---TRENDS---" and then identify 3-5 quantitative trends from the data. Format each as:
TREND: [Specific trend with numbers/percentages/timeframes]

Examples of trends to look for:
- Cost increases/decreases over time (e.g., "Dental supplies increased 5% over past 90 days")
- Payment timing changes (e.g., "Average days to payment extended from 38 to 44 days over last 12 months")
- Volume trends (e.g., "Patient visits declined 12% quarter-over-quarter")
- Revenue patterns (e.g., "Collections rate dropped from 94% to 89% in Q3")
- Expense patterns (e.g., "Lab costs up 15% year-over-year")

Be specific with numbers, percentages, and timeframes when identifying trends.

SECTION 3 - STRUCTURED INTERNAL REVIEW JSON:
After the plain-text sections, include exactly one JSON object between these markers:
{STRUCTURED_ANALYSIS_START}
{{
  "schemaVersion": "{STRUCTURED_ANALYSIS_SCHEMA_VERSION}",
  "toolType": "{normalized_tool_type}",
  "executiveSummary": {{
    "summary": "",
    "primaryConcern": "",
    "recommendedFocus": ""
  }},
  "rankedFindings": [
    {{
      "rank": 1,
      "title": "",
      "category": "",
      "severity": "low | medium | high | critical",
      "confidence": "low | medium | high",
      "evidence": [
        {{
          "label": "",
          "value": "",
          "sourceHint": ""
        }}
      ],
      "financialValue": "",
      "operationalImplication": "",
      "recommendedAction": "",
      "followUpQuestion": "",
      "implementationDifficulty": "low | medium | high",
      "estimatedImpactCategory": "cash_flow | revenue_leakage | workflow_efficiency | growth | compliance | data_quality",
      "clientFacingSummary": "",
      "internalReviewerNotes": ""
    }}
  ],
  "dataQualityNotes": [],
  "implementationPriorities": [],
  "consultantChecklist": [],
  "suggestedReportSections": []
}}
{STRUCTURED_ANALYSIS_END}

Structured JSON rules:
- The JSON must be valid JSON with double-quoted keys and strings.
- Keep evidence metric-based and concise; do not include raw extracted rows, raw document text, PHI, filenames, storage paths, signed URLs, tokens, or secrets.
- Separate client-facing wording from internal reviewer notes.
- Rank findings by operational urgency and review value."""


def _tool_specific_prompt_focus(tool_type: str) -> str:
    if tool_type == "ar":
        return """- Aging bucket risk and high-risk balances
- Payer vs patient split if available
- Collection workflow bottlenecks
- Claim follow-up priority
- Cash-flow risk
- Immediate action list
- Data limitations and missing fields"""
    if tool_type == "claims":
        return """- Denial or rejection patterns
- Payer-specific issues
- Documentation or coding issues
- Appeal opportunities
- Follow-up priorities
- Operational bottlenecks
- High-risk claims
- Data limitations and missing fields"""
    return """- Production trends
- Gross and net production
- Collections
- Writeoffs and adjustments
- Expense and cost patterns
- Profitability and cash-flow implications
- Volatility or inconsistency
- Revenue leakage
- Growth, new-patient, or membership signals if present
- Data limitations and missing fields"""


def _openai_analysis(data_input: str, *, tool_type: str = "financial") -> str:
    from openai import OpenAI

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("missing_openai_api_key")

    client = OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model=_get_model_config()["openai"],
        messages=[
            {"role": "system", "content": _get_analysis_prompt(tool_type)},
            {"role": "user", "content": f"Analyze this dental practice data:\n\n{data_input[:6000]}"},
        ],
        temperature=0.3,
        max_tokens=2500,
    )
    return response.choices[0].message.content


def _xai_analysis(data_input: str, *, tool_type: str = "financial") -> str:
    from openai import OpenAI

    api_key = os.getenv("XAI_API_KEY")
    if not api_key:
        raise RuntimeError("missing_xai_api_key")

    client = OpenAI(
        api_key=api_key,
        base_url="https://api.x.ai/v1",
    )
    response = client.chat.completions.create(
        model=_get_model_config()["xai"],
        messages=[
            {"role": "system", "content": _get_analysis_prompt(tool_type)},
            {"role": "user", "content": f"Analyze this dental practice data:\n\n{data_input[:6000]}"},
        ],
        temperature=0.3,
        max_tokens=2500,
    )
    return response.choices[0].message.content


def _anthropic_analysis(data_input: str, *, tool_type: str = "financial") -> str:
    import anthropic

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("missing_anthropic_api_key")

    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model=_get_model_config()["anthropic"],
        max_tokens=2500,
        temperature=0.3,
        system=_get_analysis_prompt(tool_type),
        messages=[
            {"role": "user", "content": f"Analyze this dental practice data:\n\n{data_input[:6000]}"}
        ],
    )
    return message.content[0].text


def _run_provider_analysis_with_retry(
    *,
    provider_name: str,
    analysis_func: Callable[[str], str],
    data_input: str,
) -> tuple[str, bool, Optional[str]]:
    last_error_type: Optional[str] = None
    for attempt in range(1, MAX_PROVIDER_RETRIES + 2):
        try:
            analysis_text = analysis_func(data_input)
            if not isinstance(analysis_text, str) or not analysis_text.strip():
                raise ValueError("empty_provider_response")
            if attempt > 1:
                logger.info(
                    "[admin_analysis] provider recovered provider=%s attempt=%s",
                    provider_name,
                    attempt,
                )
            return analysis_text, True, None
        except Exception as exc:
            transient = _is_transient_provider_error(exc)
            last_error_type = _safe_provider_error_type(exc)
            logger.warning(
                "[admin_analysis] provider failed provider=%s attempt=%s transient=%s error_type=%s",
                provider_name,
                attempt,
                transient,
                last_error_type,
            )
            if not transient or attempt > MAX_PROVIDER_RETRIES:
                break
            time.sleep(PROVIDER_RETRY_BACKOFF_SECONDS * attempt)

    logger.warning("[admin_analysis] provider unavailable provider=%s", provider_name)
    return PROVIDER_UNAVAILABLE_MESSAGE, False, last_error_type


def _is_transient_provider_error(exc: Exception) -> bool:
    status_code = _provider_status_code(exc)
    if status_code in TRANSIENT_PROVIDER_STATUS_CODES:
        return True

    safe_text = f"{type(exc).__name__} {str(exc)}".lower()
    transient_markers = (
        "rate limit",
        "rate_limit",
        "overload",
        "overloaded",
        "temporarily unavailable",
        "temporary provider capacity",
        "timeout",
        "timed out",
        "server error",
        "internal server",
        "internal_server_error",
        "service unavailable",
        "bad gateway",
        "gateway timeout",
        "connection reset",
    )
    return any(marker in safe_text for marker in transient_markers)


def _provider_status_code(exc: Exception) -> Optional[int]:
    for attr in ("status_code", "status"):
        value = getattr(exc, attr, None)
        if isinstance(value, int):
            return value

    response = getattr(exc, "response", None)
    value = getattr(response, "status_code", None)
    if isinstance(value, int):
        return value
    return None


def _safe_provider_error_type(exc: Exception) -> str:
    status_code = _provider_status_code(exc)
    if status_code is not None:
        return f"{type(exc).__name__}:{status_code}"
    return type(exc).__name__


def _raise_if_canceled(cancel_checker: Optional[CancelChecker]) -> None:
    if cancel_checker and cancel_checker():
        raise AdminFinancialProcessingCanceled()
